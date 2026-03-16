import os
import json
import requests
import subprocess
from openpyxl import Workbook, load_workbook

# ==========================================
# CONFIGURACIÓN
# ==========================================
ARCHIVO_EXCEL = "coleccion_pokemon.xlsx"
CARPETA_DATOS = "datos_local" 

DICCIONARIO_EXPANSIONES = {
    "SVI": "sv1", "PAL": "sv2", "OBF": "sv3", "MEW": "sv3pt5",
    "PAR": "sv4", "PAF": "sv4pt5", "TEF": "sv5", "TWM": "sv6",
    "SFA": "sv6pt5", "SCR": "sv7", "SSP": "sv8",
    "SSH": "swsh1", "RCL": "swsh2", "DAA": "swsh3", "CPA": "swsh3pt5",
    "VIV": "swsh4", "SHF": "swsh4pt5", "BST": "swsh5", "CRE": "swsh6",
    "EVS": "swsh7", "FST": "swsh8", "BRS": "swsh9", "ASR": "swsh10",
    "PGO": "pgo", "LOR": "swsh11", "SIT": "swsh12", "CRZ": "swsh12pt5"
}

DICCIONARIO_TIPOS = {
    "Grass": "Planta", "Fire": "Fuego", "Water": "Agua",
    "Lightning": "Eléctrico", "Psychic": "Psíquico", "Fighting": "Lucha",
    "Darkness": "Siniestro", "Metal": "Metálico", "Colorless": "Incoloro",
    "Dragon": "Dragón", "Fairy": "Hada", "Trainer": "Entrenador", "Energy": "Energía"
}

if not os.path.exists(CARPETA_DATOS):
    os.makedirs(CARPETA_DATOS)

def obtener_datos_set_local(api_set_id):
    ruta_archivo = os.path.join(CARPETA_DATOS, f"{api_set_id}.json")
    if not os.path.exists(ruta_archivo):
        print(f"📥 Descargando base de datos de '{api_set_id}' desde GitHub...")
        url_github = f"https://raw.githubusercontent.com/PokemonTCG/pokemon-tcg-data/master/cards/en/{api_set_id}.json"
        try:
            respuesta = requests.get(url_github, timeout=15)
            respuesta.raise_for_status()
            with open(ruta_archivo, 'w', encoding='utf-8') as f:
                f.write(respuesta.text)
            print("✅ Expansión guardada en local correctamente.")
        except Exception as e:
            print(f"❌ Error al descargar de GitHub: {e}")
            return None
    try:
        with open(ruta_archivo, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"❌ Error al leer el archivo local: {e}")
        return None

def buscar_carta_local(siglas, numero):
    api_set_id = DICCIONARIO_EXPANSIONES.get(siglas.upper())
    if not api_set_id:
        print(f"❌ Error: Las siglas '{siglas}' no están registradas.")
        return None

    cartas_set = obtener_datos_set_local(api_set_id)
    if not cartas_set: return None
        
    print(f"🔍 Buscando la carta {numero} en modo local...")
    
    carta_encontrada = None
    for carta in cartas_set:
        num_json = str(carta.get("number", "")).lstrip("0")
        num_buscado = str(numero).lstrip("0")
        if num_json == num_buscado:
            carta_encontrada = carta
            break
            
    if carta_encontrada:
        nombre = carta_encontrada.get('name', 'Desconocido')
        supertype = carta_encontrada.get('supertype', 'Desconocido')
        if supertype == "Pokémon":
            tipos = carta_encontrada.get('types', ['Colorless'])
            tipo_ingles = tipos[0]
        else:
            tipo_ingles = supertype
            
        tipo_espanol = DICCIONARIO_TIPOS.get(tipo_ingles, tipo_ingles)
            
        return {
            "nombre": nombre, "tipo": tipo_espanol,
            "siglas": siglas.upper(), "numero": carta_encontrada.get("number", str(numero))
        }
    else:
        print("❌ Carta no encontrada.")
        return None

def guardar_en_excel(datos_carta):
    if os.path.exists(ARCHIVO_EXCEL): wb = load_workbook(ARCHIVO_EXCEL)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
            
    nombre_hoja = datos_carta["tipo"]
    
    if nombre_hoja not in wb.sheetnames:
        ws = wb.create_sheet(title=nombre_hoja)
        ws.append(["Cantidad", "Nombre", "Expansión", "Número"])
    else: ws = wb[nombre_hoja]

    carta_encontrada = False
    for fila in range(2, ws.max_row + 1):
        celda_siglas = str(ws.cell(row=fila, column=3).value)
        celda_numero = str(ws.cell(row=fila, column=4).value)
        if celda_siglas == datos_carta["siglas"] and celda_numero == datos_carta["numero"]:
            cantidad_actual = ws.cell(row=fila, column=1).value
            nueva_cantidad = (cantidad_actual if cantidad_actual else 0) + 1
            ws.cell(row=fila, column=1).value = nueva_cantidad
            carta_encontrada = True
            print(f"💾 ¡Actualizada! {datos_carta['nombre']} ahora tiene cantidad: {nueva_cantidad}")
            break
            
    if not carta_encontrada:
        ws.append([1, datos_carta["nombre"], datos_carta["siglas"], datos_carta["numero"]])
        print(f"💾 ¡Nueva carta añadida a la pestaña '{nombre_hoja}': {datos_carta['nombre']}")

    wb.save(ARCHIVO_EXCEL)

def subir_a_github():
    print("☁️ Comprobando si hay cambios para subir a GitHub...")
    try:
        # Añadir el Excel a los cambios preparados
        subprocess.run(["git", "add", ARCHIVO_EXCEL], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        
        # Intentar hacer commit. Si no hay cambios, git devuelve un error (que capturamos para que no pete el programa)
        resultado = subprocess.run(["git", "commit", "-m", "Actualización automática de colección"], capture_output=True, text=True)
        
        # Si el commit ha detectado cambios, hacemos push
        if "nothing to commit" not in resultado.stdout and "nada para hacer commit" not in resultado.stdout:
            print("🚀 Subiendo tu colección actualizada a la nube...")
            subprocess.run(["git", "push"], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            print("✅ ¡Copia de seguridad subida a GitHub con éxito!")
        else:
            print("✅ Tu colección ya está sincronizada y al día en la nube. No hay cambios nuevos.")
            
    except subprocess.CalledProcessError:
        print("⚠️ Hubo un problema al subir a GitHub. Comprueba tu conexión a internet o tus credenciales.")
    except FileNotFoundError:
        print("⚠️ No se encontró 'git'. Asegúrate de tenerlo instalado.")

def main():
    print("--- GESTOR DE CARTAS POKÉMON (MODO LOCAL + AUTO-GITHUB) ---")
    print("Escribe 'salir' en cualquier momento para cerrar y guardar en la nube.\n")
    
    while True:
        entrada = input("Introduce las siglas de la expansión y el número (ej: MEW 15): ")
        
        if entrada.lower() == 'salir':
            subir_a_github()
            print("¡Hasta la próxima!")
            break
            
        partes = entrada.strip().split()
        if len(partes) != 2:
            print("⚠️ Formato incorrecto. Ej: MEW 15")
            continue
            
        siglas = partes[0]
        numero = partes[1]
        
        datos = buscar_carta_local(siglas, numero)
        if datos:
            guardar_en_excel(datos)
        print("-" * 30)

if __name__ == "__main__":
    main()